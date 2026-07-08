import json
import time
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_ROOT / "workbook" / "Pokemon Battle Compass v1.2.xlsx"
DATA_DIR = PROJECT_ROOT / "data"

EXPORTS = [
    ("Team Data", "team_data.json", "Team Data"),
    ("Movelist", "moves.json", "Moves"),
    ("Ability Rules", "ability_rules.json", "Ability Rules"),
    ("Items", "items.json", "Items"),
    ("Opponent", "opponents.json", "Opponents"),
]


def sheet_to_records(workbook, sheet_name):
    sheet = workbook[sheet_name]

    headers = [
        sheet.cell(row=1, column=col).value
        for col in range(1, sheet.max_column + 1)
    ]

    records = []

    for row in range(2, sheet.max_row + 1):
        record = {}

        for col, header in enumerate(headers, start=1):
            record[header] = sheet.cell(row=row, column=col).value

        if any(value is not None for value in record.values()):
            records.append(record)

    return records


def export_json(records, output_file):
    output_path = DATA_DIR / output_file

    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(records, file, indent=2)

    return len(records)


def print_header():
    print("=" * 52)
    print("        Pokémon Battle Compass Import")
    print("=" * 52)
    print()
    print("Workbook")
    print("--------")
    print(WORKBOOK.name)
    print()


def print_summary(results, elapsed_seconds):
    print("Imported")
    print("--------")

    for label, count in results:
        print(f"✓ {label:<16} {count:>6} records")

    print()
    print(f"Completed successfully in {elapsed_seconds:.2f} seconds.")
    print()
    print("JSON files written to:")
    print(DATA_DIR)
    print()
    print("Ready to launch:")
    print("streamlit run app.py")


def main():
    start_time = time.perf_counter()

    print_header()

    wb = load_workbook(WORKBOOK, data_only=True)

    results = []

    for sheet_name, output_file, label in EXPORTS:
        records = sheet_to_records(wb, sheet_name)
        count = export_json(records, output_file)
        results.append((label, count))

    elapsed_seconds = time.perf_counter() - start_time

    print_summary(results, elapsed_seconds)


if __name__ == "__main__":
    main()